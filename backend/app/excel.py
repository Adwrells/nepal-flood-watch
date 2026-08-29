"""Workbook export -- rewritten in full on every cycle.

Six sheets: Dashboard (KPIs + hotlist), Stations (the scored table), Rainfall,
Incidents, News, Method. Written to a temp file then swapped into place so a
cycle can never leave a half-written workbook for someone who has it open.
"""
import json
import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .config import settings
from .scoring import BANDS

HEADER_FILL = PatternFill("solid", fgColor="0F172A")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
TITLE_FONT = Font(bold=True, size=14, color="0F172A")

# Same five bands the UI uses, as Excel fills.
BAND_FILL = {
    "SEVERE": "D946EF",
    "DANGER": "F43F5E",
    "WARNING": "FB923C",
    "WATCH": "FDE047",
    "NORMAL": "34D399",
}


def _sheet(wb, title, headers, rows, widths=None):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for cell in ws[1]:
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    for r in rows:
        ws.append(r)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for i, w in enumerate(widths or [], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return ws


def export(stations, scores, rain_by_id, incidents, news) -> None:
    by_id = {s["id"]: s for s in stations}
    scores = sorted(scores, key=lambda s: s["fsi"], reverse=True)
    now = datetime.now().astimezone().strftime("%Y-%m-%d %H:%M %Z")

    wb = Workbook()
    wb.remove(wb.active)

    # --- Dashboard --------------------------------------------------------
    ws = wb.create_sheet("Dashboard")
    ws["A1"], ws["A1"].font = "Nepal Flood Watch - Live Severity", TITLE_FONT
    ws["A2"] = f"Generated {now}  |  refresh every {settings.cycle_minutes} min"
    ws["A2"].font = Font(italic=True, size=9, color="475569")

    counts = {label: sum(1 for s in scores if s["band"] == label) for _, label in BANDS}
    ws.append([]), ws.append(["Band", "Stations", "Meaning"])
    for cell in ws[4]:
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT
    meanings = {
        "SEVERE": "At or beyond danger mark and still rising",
        "DANGER": "Danger mark reached or imminent",
        "WARNING": "Between warning and danger mark",
        "WATCH": "Elevated - rising or heavy rain upstream",
        "NORMAL": "Within normal range",
    }
    for _, label in BANDS:
        ws.append([label, counts[label], meanings[label]])
        ws.cell(ws.max_row, 1).fill = PatternFill("solid", fgColor=BAND_FILL[label])

    ws.append([]), ws.append(["Top 25 stations by severity"])
    ws.cell(ws.max_row, 1).font = Font(bold=True, size=12)
    head = ws.max_row + 1
    ws.append(["Station", "District", "Basin", "FSI", "Band", "P(danger in 6h)", "Level m", "Danger m"])
    for cell in ws[head]:
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT
    for s in scores[:25]:
        st = by_id.get(s["station_id"], {})
        ws.append([
            st.get("name"), st.get("district"), st.get("basin"), s["fsi"], s["band"],
            s["p_exceed_6h"], st.get("level"), st.get("danger_level"),
        ])
        ws.cell(ws.max_row, 5).fill = PatternFill("solid", fgColor=BAND_FILL[s["band"]])
        ws.cell(ws.max_row, 6).number_format = "0.0%"
    for col, w in zip("ABCDEFGH", [42, 16, 18, 8, 11, 16, 10, 10]):
        ws.column_dimensions[col].width = w

    # --- Stations ---------------------------------------------------------
    rows = []
    for s in scores:
        st, comp = by_id.get(s["station_id"], {}), json.loads(s["components"])
        rain = rain_by_id.get(s["station_id"], {})
        rows.append([
            s["station_id"], st.get("name"), st.get("district"), st.get("basin"),
            st.get("lat"), st.get("lon"), st.get("level"), st.get("warning_level"),
            st.get("danger_level"), s["rise_rate"], rain.get("past_24h"), rain.get("next_12h"),
            comp["level"], comp["rise"], comp["rain"], comp["corroboration"],
            s["fsi"], s["band"], s["p_exceed_6h"], st.get("steady"), st.get("ts"),
        ])
    ws = _sheet(
        wb, "Stations",
        ["ID", "Station", "District", "Basin", "Lat", "Lon", "Level m", "Warning m",
         "Danger m", "Rise m/h", "Rain 24h mm", "Rain +12h mm", "C:Level", "C:Rise",
         "C:Rain", "C:Corrob", "FSI", "Band", "P(6h)", "Trend", "Reading time"],
        rows,
        [7, 40, 15, 16, 9, 9, 9, 10, 9, 9, 12, 13, 9, 9, 9, 10, 8, 11, 8, 10, 22],
    )
    # Colour the FSI column by band threshold, high to low so each rule wins its range.
    for floor, label in BANDS:
        ws.conditional_formatting.add(
            f"Q2:Q{ws.max_row}",
            CellIsRule(operator="greaterThanOrEqual", formula=[str(floor)],
                       fill=PatternFill("solid", fgColor=BAND_FILL[label])),
        )
    for r in range(2, ws.max_row + 1):
        ws.cell(r, 19).number_format = "0.0%"

    # --- Supporting sheets -------------------------------------------------
    _sheet(wb, "Rainfall",
           ["Station ID", "Station", "Past 24h mm", "Next 12h mm", "As of"],
           [[r["station_id"], by_id.get(r["station_id"], {}).get("name"),
             r["past_24h"], r["next_12h"], r["ts"]] for r in rain_by_id.values()],
           [11, 40, 13, 13, 20])

    _sheet(wb, "Incidents",
           ["Date", "Title", "Hazard", "Lat", "Lon", "Source", "URL"],
           [[i["occurred_on"], i["title"], i["hazard"], i["lat"], i["lon"], i["source"], i["url"]]
            for i in incidents],
           [22, 52, 14, 10, 10, 22, 46])

    _sheet(wb, "News",
           ["Published", "Headline", "Districts", "Source", "URL"],
           [[n["published"], n["title"], n["districts"], n["source"], n["url"]] for n in news],
           [30, 62, 20, 18, 46])

    # --- Method ------------------------------------------------------------
    ws = wb.create_sheet("Method")
    ws["A1"], ws["A1"].font = "How the Flood Severity Index is built", TITLE_FONT
    for row in [
        [],
        ["Component", "Weight", "Full scale", "Source"],
        ["Level vs danger mark", settings.w_level, "gauge at danger mark = 85-100", "DHM river watch"],
        ["Rate of rise", settings.w_rise, "0.50 m/h = 100", "DHM, cycle-over-cycle delta"],
        ["Rainfall pressure", settings.w_rain, "200 mm (24h past + 12h fcst) = 100", "Open-Meteo"],
        ["Corroboration", settings.w_corroboration, "incident within 25 km = 40", "BIPAD + news RSS"],
        [],
        ["P(danger in 6h) = logistic((level + 6*rise + rain push - danger) / 0.35)"],
        ["Bands: SEVERE 90+, DANGER 75-89, WARNING 50-74, WATCH 25-49, NORMAL 0-24"],
        [],
        ["Not an official warning. DHM and MoHA issue the authoritative alerts."],
    ]:
        ws.append(row)
    for cell in ws[3]:
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT
    for col, w in zip("ABCD", [26, 10, 34, 30]):
        ws.column_dimensions[col].width = w

    # Atomic swap so an open workbook never sees a partial write.
    tmp = settings.excel_path.with_suffix(".tmp.xlsx")
    wb.save(tmp)
    os.replace(tmp, settings.excel_path)
